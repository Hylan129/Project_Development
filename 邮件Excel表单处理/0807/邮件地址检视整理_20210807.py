#!/usr/bin/python
# coding=UTF-8

import time,tkinter,os,xlrd,xlwt,openpyxl
#from tkinter.constants import CENTER
#from xlrd.formula import _TOKEN_NOT_ALLOWED
from tkinter import ttk
from tkinter.filedialog import *
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font


global data_raw,alfa,mail_col_num,cus_col_num,data_row_num
alfa = {}
empty_data = []
mail_col_num = 1
cus_col_num = 0
data_row_num = 0

#定义主界面：
root  = tkinter.Tk()
root.minsize(500,600)
root.maxsize(500,600)

root.title('邮件地址表单辅助处理-20210807_By Hylan')

command1 = tkinter.StringVar()
command2 = tkinter.StringVar()
dirname = tkinter.StringVar()
filenames = tkinter.StringVar()
filename1 = tkinter.StringVar()
filename2 = tkinter.StringVar()

address = tkinter.StringVar()
customer_name = tkinter.StringVar()

Choice1 = tkinter.StringVar()

button1 = tkinter.Button(root,text='退订查验处理',command = lambda : update_data_deal())
button1.place(x=180,y=180,width=129,height=40)

button2 = tkinter.Button(root,text='新增查验处理',command = lambda : new_data_deal())
button2.place(x=340,y=180,width=129,height=40)

button3 = tkinter.Button(root,text='合并去重大整理',command = lambda : big_data_deal_dir())
button3.place(x=27,y=180,width=129,height=40)

# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='F1文件目录：').place(x=27,y=240,width=129,height=40)
tkinter.Entry(root, textvariable=dirname,state='disabled').place(x=150,y=240,width=240,height=40)
tkinter.Button(root, text='选择目录', command=lambda:select_Dir()).place(x=400,y=240,width=70,height=40)

# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='F2文件清单：').place(x=27,y=280,width=129,height=40)
tkinter.Entry(root, textvariable=filenames).place(x=150,y=280,width=240,height=40)
button4 = tkinter.Button(root, text='选择文件', command=lambda:big_data_deal_file())
button4.place(x=400,y=280,width=70,height=40)


# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='退订清单查验：').place(x=27,y=320,width=129,height=40)
tkinter.Entry(root, textvariable=filename1).place(x=150,y=320,width=240,height=40)
button5 = tkinter.Button(root, text='选择文件', command=lambda:selectFiles_Update())
button5.place(x=400,y=320,width=70,height=40)


# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='新客户清单查验：').place(x=27,y=360,width=129,height=40)
entry6 = tkinter.Entry(root, textvariable=filename2)
entry6.place(x=150,y=360,width=240,height=40)
button6 = tkinter.Button(root, text='选择文件', command=lambda:selectFiles_NEW())
button6.place(x=400,y=360,width=70,height=40)

label_newfunction = tkinter.Label(root,text="* 数据检索查询：")
label_newfunction.place(x=30,y=420)
# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='#地址查询：').place(x=27,y=460,width=129,height=40)
Entry_CheckMail = tkinter.Entry(root, textvariable=address)
Entry_CheckMail.place(x=150,y=460,width=240,height=40)
tkinter.Button(root, text='点击查询', command=lambda:checkemail()).place(x=400,y=460,width=70,height=40)

# 构建“选择文件”这一行的标签、输入框以及启动按钮，同时我们希望当用户选择图片之后能够显示原图的基本信息
tkinter.Label(root, text='#查询结果：').place(x=27,y=500,width=129,height=40)
result_display = tkinter.Text(root,bg='#F2F2EB',selectforeground='red',width=45,height=5)
result_display.place(x=150,y=510)

label1 = tkinter.Label(root,text="* 动态进展：")
label1.place(x=30,y=5)

text = tkinter.Text(root,bd = 0.5,bg='#F2F2EB',selectforeground='yellow',width=63,height=9)
scroll = tkinter.Scrollbar()
# 放到窗口的右侧, 填充Y竖直方向
scroll.pack(side=tkinter.RIGHT,fill=tkinter.Y)
 
# 两个控件关联
scroll.config(command=text.yview)
text.config(yscrollcommand=scroll.set)
text.place(x=29,y=25)

#excel style：
style1 = xlwt.easyxf("font: height 300 ,name Times New Roman,color-index red, bold on; align: wrap on,vert centre,horiz centre;")
style2 = xlwt.easyxf("align: wrap on,vert centre,horiz left;")
style3 = xlwt.easyxf("font: bold on;align: wrap on,vert centre,horiz left;")

def write_excel_1(content_list_all):
	try:
		#excel 写入
		myWorkbook = xlwt.Workbook()
		mySheet = myWorkbook.add_sheet('result') # 添加活页博
		#数据写入，写入标题
		mySheet.write_merge(0, 1, 0,8,"邮件地址数据查重整理报告清单_存在空值",style1) 

		row_marks = ['No.','客户邮件地址','客户名称','所在文件路径','所在活页簿名称','所在行号','其他1','其他2','其他3']
		
		for num,row_mark in enumerate(row_marks):
			mySheet.write(3,num,row_mark,style3)

		#数据循环写入
		for i,contents in enumerate(content_list_all): #content_list_all为需写入的数组数据
			mySheet.write(i+4,0,i+1,style2)
			for j,content in enumerate(contents):
				mySheet.write(i+4,j+1,content,style2)
		
		col_width = [3,16,10,30,8,10,10,10,10]
		# 设置自适应列宽
		for i in range(0, len(row_marks)):
			# 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
			mySheet.col(i).width = 256 * (col_width[i]) *2

		myWorkbook.save('邮件地址查重整理清单_异常_' + time.strftime("%Y-%m-%d_%H-%M-%S")+'.xlsx') #保存excle数据表。
	except Exception as e:
		showinfo("程序报错：write_excel_1 "+str(e))

def write_excel_2(content_list_all):
	try:
		# 按照重复次数排序
		content_list_all = sorted(content_list_all.items(),key = lambda x:x[1][0],reverse=True)
		special_num1 = special_num2 = -1
		for jjj,kkk in enumerate(content_list_all):
			if kkk[1][0] != 1:
				continue
			else:
				special_num2 = jjj
				break
		if special_num2 != special_num1:
			content_list_all = content_list_all[0:special_num2] + sorted(content_list_all[special_num2:],key = lambda x:x[0].strip().split('.')[-1],reverse=False)
			
		#excel 写入
		myWorkbook = Workbook()
		mySheet = myWorkbook.get_sheet_by_name('Sheet')# 添加活页博
		mySheet.title = "result"
		base_style = Alignment(wrapText=True,vertical = 'center',horizontal="left")
		mySheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=10)
		
		mySheet.cell(1,1).value = "邮件地址数据查重整理报告清单"
		mySheet.cell(1,1).alignment = Alignment(wrapText=True,vertical = 'center',horizontal="center")
		mySheet.cell(1,1).font = Font(bold=True, color="00FF0000",size=18)
		#设定列宽
		col_width = [3,16,5,32,8,6,10,10,10,10]
		col_num = ["A","B","C","D","E","F","G","H","I","J"]
		for num in range(10):
			mySheet.column_dimensions[col_num[num]].width = col_width[num] * 2
			
		#隔行写入标题
		row_marks = ['No.','客户邮件地址','存在个数','所在文件路径','所在活页簿名称','所在行号','客户名称','其他1','其他2','其他3']
		mySheet.cell(3,1).value = ''
		mySheet.append(row_marks)
		for num_cell in range(10):
			mySheet.cell(4,num_cell+1).alignment = base_style
			mySheet.cell(4,num_cell+1).font = Font(size=11,bold=True)
		
		#数据循环写入
		for i,contents in enumerate(content_list_all): #content_list_all为需写入的数组数据
			new_contents = [contents[0],contents[1][0]]
			new_contents_back = ['','','','']
			for mn in range(1,len(contents[1])):
				for j in range(4):
					new_contents_back[j] +=  str(contents[1][mn][j]) + "\r\n" 
			
			new_contents_back[3] = '\r\n'.join(list(set(new_contents_back[3].strip().split("\r\n"))))
			
			new_contents += [iii.strip() for iii in new_contents_back]
			new_contents += contents[1][1][4:]
			try:
				mySheet.append([i+1]+new_contents)
			except Exception as e:
				showinfo("数据异常：write_excel_2  error："+str(e) + "-->异常数据：" + '行号：'+str(i+1)+" 邮件地址："+ new_contents[0])
				with open('异常字符数据集_待处理.txt','a') as code:
					code.write("\n>>【异常信息-开始】\n")
					code.write('-'.join([str(jjj) for jjj in [i+1]+new_contents]))
					code.write("\n【异常信息-结束】\n")
			for num_cell in range(10):
				mySheet.cell(5+i,num_cell+1).alignment = base_style
				mySheet.cell(5+i,num_cell+1).font = Font(size=10)

		myWorkbook.save('邮件地址查重整理清单_正常_' + time.strftime("%Y-%m-%d_%H-%M-%S")+'.xlsx') #保存excle数据表。
	except Exception as e:
		showinfo("程序报错：write_excel_2 "+str(e))

def write_excel_3(content_list_all,row_marks,title_name):
	try:
		#excel 写入
		myWorkbook = xlwt.Workbook()
		mySheet = myWorkbook.add_sheet('result') # 添加活页博
		#数据写入，写入标题
		mySheet.write_merge(0, 1, 0,len(row_marks)-1,title_name,style1) 
		
		for num,row_mark in enumerate(row_marks):
			mySheet.write(3,num,row_mark,style3)

		#数据循环写入
		for i,contents in enumerate(content_list_all): #content_list_all为需写入的数组数据
			mySheet.write(i+4,0,i+1,style2)
			#print(i,contents)
			for j,content in enumerate(contents):
				mySheet.write(i+4,j+1,content,style2)
		
		col_width = [3,16,10,5,30,8,6,10,10,10,10]
		# 设置自适应列宽
		for i in range(0, len(row_marks)):
			# 256*字符数得到excel列宽,为了不显得特别紧凑添加两个字符宽度
			mySheet.col(i).width = 256 * (col_width[i]) *2

		myWorkbook.save(title_name + "_"+ time.strftime("%Y-%m-%d_%H-%M-%S")+'.xlsx') #保存excle数据表。
	except Exception as e:
		showinfo("程序报错：write_excel_3 "+str(e))


def checkemail(mail=None):
	mail_need_check = address.get()
	if not mail_need_check.strip():
		return feedback('异常报错','输入查询地址为空。') 
	if "##" in mail_need_check:
		global mail_col_num,cus_col_num,data_row_num
		mail_col_num,cus_col_num,data_row_num = [int(num) for num in mail_need_check.strip().split('##')]
		showinfo("参数重置：" + "mail_col_num、cus_col_num、data_row_num重置为" + str(mail_col_num) + "、" + str(cus_col_num)+"、"+str(data_row_num))
	try:
		result = alfa[mail_need_check.strip()]
		result_display.delete('1.0','end')
		result_display.insert('1.0',str(result[0])+"个："+'#'.join(['-->'.join(j) for j in result[1:]]))
		showinfo("查询数据："+ mail_need_check + "共"+str(result[0])+'个。')
	except :
		showinfo("查询数据："+ mail_need_check + "未查询到！")

def selectFiles_Update():

	filepaths = askopenfilename(filetypes=[("office 2007 excel files","*.xls"),(" office 2010 excel files ","*.xlsx"),\
                                ("All files","*.*")], title="Select files")  # 选择打开什么文件，返回文件名
	filename1.set(filepaths)            # 设置变量filename的值
	showinfo('文件选择：'+'退订文件清单选择成功！'+filepaths)
	#button5.config(state='disabled')
	"""
	data_deal(filepaths)
	button5.config(state='normal')
	"""

def selectFiles_NEW():
	filepaths = askopenfilename(filetypes=[("office 2007 excel files","*.xls"),(" office 2010 excel files ","*.xlsx"),\
                                ("All files","*.*")], title="Select files")  # 选择打开什么文件，返回文件名
	filename2.set(filepaths)             # 设置变量filename的值
	showinfo('文件选择：'+'新增客户文件清单选择成功！'+filepaths)
	#button6.config(state='disabled')
	"""
	data_deal(filepaths)
	button6.config(state='normal')
	"""

def update_data_deal():
	update_file_dir,total_file_dir = filename1.get(),dirname.get()

	if not update_file_dir.strip():
		return feedback('异常报错','退订清单文件选择为空。')
	if not total_file_dir.strip():
		return feedback('异常报错','F1清单文件选择为空。')
	#print(update_file_dir,total_file_dir)
	update_data, total_data = data_deal_return([update_file_dir],mail_col_num,cus_col_num,data_row_num),data_deal_return(getDir_list(total_file_dir),mail_col_num,cus_col_num,data_row_num)
	update_data_mails_ok,update_data_mails_ng = update_data
	total_data_mails_ok,total_data_mails_ng = total_data

	if update_data_mails_ng != []:
		showinfo("数据告警："+"退订清单中存在空值条目，请自行检查！")
	for update_date_mail in update_data_mails_ok.keys():
		if update_date_mail in total_data_mails_ok.keys():
			update_data_mails_ok[update_date_mail] = [update_date_mail,update_data_mails_ok[update_date_mail][1][-4],'是']+total_data_mails_ok[update_date_mail][1][0:4] + update_data_mails_ok[update_date_mail][1][4:]
		else:
			update_data_mails_ok[update_date_mail] = [update_date_mail,update_data_mails_ok[update_date_mail][1][-4], '否']
	
	row_marks = ['No.','客户邮件地址','客户名称1','是否订阅过','所在文件路径','所在活页簿名称','所在行号','客户名称2','其他1','其他2','其他3']
	title_name = "退订邮件地址查验整理报告"
	write_excel_3(update_data_mails_ok.values(),row_marks,title_name)
	showinfo("结果输出：Done！退订清单查验完成！")

	button5.config(state='normal')

def new_data_deal():
	new_file_dir,total_file_dir = filename2.get(),dirname.get()

	if not new_file_dir.strip():
		return feedback('异常报错','新增客户清单文件选择为空。')
	if not total_file_dir.strip():
		return feedback('异常报错','F1清单文件选择为空。')
	#print(new_file_dir,total_file_dir)
	new_data, total_data = data_deal_return([new_file_dir],mail_col_num,cus_col_num,data_row_num),data_deal_return(getDir_list(total_file_dir),mail_col_num,cus_col_num,data_row_num)
	new_data_mails_ok,new_data_mails_ng = new_data
	total_data_mails_ok,total_data_mails_ng = total_data

	if new_data_mails_ng != []:
		showinfo("数据告警："+"新增客户清单中存在空值条目，请自行检查！")

	for new_date_mail in new_data_mails_ok.keys():
		if new_date_mail in total_data_mails_ok.keys():
			new_data_mails_ok[new_date_mail] = [new_date_mail,new_data_mails_ok[new_date_mail][1][-4],'是']+total_data_mails_ok[new_date_mail][1][0:4] + new_data_mails_ok[new_date_mail][1][4:]
			#print(new_data_mails_ok)
		else:
			new_data_mails_ok[new_date_mail] = [new_date_mail,new_data_mails_ok[new_date_mail][1][-4], '否'] + ["","","",""] + new_data_mails_ok[new_date_mail][1][4:]
	
	row_marks = ['No.','客户邮件地址','客户名称1','是否在已有清单中','所在文件路径','所在活页簿名称','所在行号','客户名称2','其他1','其他2','其他3']
	title_name = "新增客户邮件地址查验整理报告"
	write_excel_3(new_data_mails_ok.values(),row_marks,title_name)
	showinfo("结果输出：Done！新增客户邮件清单查验完成！")

	button6.config(state='normal')

def getDir_list(path):
	paths = []
	if os.path.isdir(path):
		for dir,zi_dirs,documents in os.walk(path):
			for document in documents:
				if document.split('.')[-1] in ['xls','xlsx']:
					paths.append(dir + '/'+document)
		return paths
	return ''

def showinfo(result="功能测试：test ok！"):
	text.config(state='normal')
	realtime = time.strftime("%Y-%m-%d %H:%M:%S ")
	textvar = realtime + result #系统时间和传入结果
	text.insert('1.0',textvar) #显示在text框里面
	text.insert('1.0','\n') #换行
	text.tag_add('content1','2.10','2.20')
	text.tag_add('content2','2.20','2.24')
	text.tag_config('content1',foreground='blue')
	text.tag_config('content2',foreground='red')
	text.see('1.0')
	#text.config(state=DISABLED)
	text.update()

def get_xlsdata(path):
    try:
        data = xlrd.open_workbook(path)#打开需要读取的excel表
        return data
    except Exception as e:
        with open('error.txt','a') as code:
            code.write("get_xlsdata()出错：" + str(e)+"\n\n")
        showinfo("程序报错: get_xlsdata " + str(e))
        return ''
	
def select_Dir():
	#filepath = askopenfilename()  # 选择打开什么文件，返回文件名
	#filepath_dir = askopenfilenames()
	filepath_dir = askdirectory()
	dirname.set(filepath_dir)             # 设置变量filename的值

def big_data_deal_dir():
	filepath_dir = dirname.get()
	if not filepath_dir.strip():
		return feedback('异常报错','目录选择为空。')
	button3.config(state='disabled')
	data_deal(getDir_list(filepath_dir),mail_col_num,cus_col_num,data_row_num)
	button3.config(state='normal')

#表格数据处理后返回数据
def data_deal_return(files,mail_col,cusname_col,data_row):
	normal_data ={}
	unormal_data =[]
	try:
		count_file = count_huoye = mail_ok = mail_ng = 0
		for document in files:
			data_raw = get_xlsdata(document)
			if data_raw =='':
				showinfo("读取文件："+document+"异常；请重新选择数据！")
			else:
				showinfo("读取文件："+document+"正常；数据读取成功！")
			all_sheets = data_raw.sheet_names()
			try:
				for sheet in all_sheets:
					huoyebo = data_raw.sheet_by_name(sheet)
					showinfo("数据展示：活页簿 "+sheet + " 列数："+str(huoyebo.ncols)+"，行数："+str(huoyebo.nrows) + " -->："+ document)
					if (huoyebo.ncols >= 5):
						for row,content_information in enumerate(zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],huoyebo.col_values(mail_col+2)[data_row:],huoyebo.col_values(mail_col+3)[data_row:])):
							try:
								mail_addr,customer_name = content_information[0:2]
								if str(mail_addr).strip() and str(customer_name).strip() and str(mail_addr).count("@") ==1:
									if mail_addr.upper() not in normal_data.keys():
										normal_data[mail_addr.upper()]=[1,[document,sheet,str(row + 1),customer_name]+list(content_information[2:])]
									else:
										normal_data[mail_addr.upper()].append([document,sheet,str(row+1),customer_name]+list(content_information[2:]))
										normal_data[mail_addr.upper()][0] += 1
									#可视化处理：
									#showinfo("数据整理：" + mail_addr+"地址发现" + str(alfa[mail_addr][0]) + "个")
									mail_ok += 1
								else:
									unormal_data.append([mail_addr,customer_name,document,sheet,str(row + 1)]+list(content_information[2:]))
									mail_ng += 1
							except Exception as e:
								showinfo("数据整理：整理异常，异常文件地址："+document+"-->"+sheet+"-->"+str(row) + "-->"+str(mail_addr)+"-->"+str(customer_name)+" 。报错信息："+ str(e))
					else:
						blank_data = ['' for n in range(huoyebo.nrows)]
						if huoyebo.ncols == 4:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],huoyebo.col_values(mail_col+2)[data_row:],blank_data)
						elif huoyebo.ncols == 3:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],blank_data,blank_data)
						elif huoyebo.ncols == 2:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],blank_data,blank_data,blank_data)
							
						for row,content_information in enumerate(qingdan_data):
							try:
								mail_addr,customer_name = content_information[0:2]
								if str(mail_addr).strip() and str(customer_name).strip() and str(mail_addr).count("@") ==1:
									if mail_addr.upper() not in normal_data.keys():
										normal_data[mail_addr.upper()]=[1,[document,sheet,str(row + 1),str(customer_name).strip()]+list(content_information[2:])]
									else:
										normal_data[mail_addr.upper()].append([document,sheet,str(row+1),str(customer_name).strip()]+list(content_information[2:]))
										normal_data[mail_addr.upper()][0] += 1
									#可视化处理：
									#showinfo("数据整理：" + mail_addr+"地址发现" + str(alfa[mail_addr][0]) + "个")
									mail_ok += 1
								else:
									unormal_data.append([mail_addr,str(customer_name).strip(),document,sheet,str(row + 1)]+list(content_information[2:]))
									mail_ng += 1
							except Exception as e:
								showinfo("数据整理：整理异常，异常文件地址："+document+"-->"+sheet+"-->"+str(row) + "-->"+str(mail_addr)+"-->"+str(customer_name)+" 。报错信息："+ str(e))
					
					count_huoye += 1
			except Exception as e:
				showinfo("数据整理：整理异常，异常文件路径："+document+"-->"+sheet+ " 。报错信息："+ str(e))
			count_file += 1
	except Exception as e:
		showinfo("程序报错：data_deal_return  "+str(e))
	return normal_data,unormal_data

#表格数据处理
def data_deal(files,mail_col,cusname_col,data_row):
	global alfa,empty_data
	alfa = {}
	empty_data = []
	try:
		count_file = count_huoye = mail_ok = mail_ng = 0
		for document in files:
			data_raw = get_xlsdata(document)
			if data_raw =='':
				showinfo("读取文件："+document+"异常；请重新选择数据！")
			else:
				showinfo("读取文件："+document+"正常；数据读取成功！")
			all_sheets = data_raw.sheet_names()
			try:
				for sheet in all_sheets:
					huoyebo = data_raw.sheet_by_name(sheet)
					showinfo("数据展示：活页簿 "+sheet + " 列数："+str(huoyebo.ncols)+"，行数："+str(huoyebo.nrows) + " -->："+ document)
					if huoyebo.ncols >=5:
						for row,content_information in enumerate(zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],huoyebo.col_values(mail_col+2)[data_row:],huoyebo.col_values(mail_col+3)[data_row:])):
							#showinfo("数据读取：数据读取正常！data_deal for循环。")
							try:
								mail_addr,customer_name = content_information[0:2]
								if str(mail_addr).strip() and str(customer_name).strip() and str(mail_addr).count("@") ==1:
									if mail_addr.upper() not in alfa.keys():
										alfa[mail_addr.upper()]=[1,[document,sheet,str(row + 1),customer_name]+list(content_information[2:])]
									else:
										alfa[mail_addr.upper()].append([document,sheet,str(row+1),customer_name]+list(content_information[2:]))
										alfa[mail_addr.upper()][0] += 1
									#可视化处理：
									#showinfo("数据整理：" + mail_addr+"地址发现" + str(alfa[mail_addr][0]) + "个")
									mail_ok += 1
								else:
									empty_data.append([mail_addr,customer_name,document,sheet,str(row + 1)]+list(content_information[2:]))
									mail_ng += 1
							except Exception as e:
								showinfo("数据整理：整理异常，异常文件地址："+document+"-->"+sheet+"-->"+str(row) + "-->"+str(mail_addr)+"-->"+str(customer_name)+" 。报错信息："+ str(e))
					else:
						blank_data = ['' for n in range(huoyebo.nrows)]
						if huoyebo.ncols == 4:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],huoyebo.col_values(mail_col+2)[data_row:],blank_data)
						elif huoyebo.ncols == 3:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],huoyebo.col_values(mail_col+1)[data_row:],blank_data,blank_data)
						elif huoyebo.ncols == 2:
							qingdan_data = zip(huoyebo.col_values(mail_col)[data_row:],huoyebo.col_values(cusname_col)[data_row:],blank_data,blank_data,blank_data)
							
						for row,content_information in enumerate(qingdan_data):
							#showinfo("数据读取：数据读取正常！data_deal for循环。")
							try:
								mail_addr,customer_name = content_information[0:2]
								if str(mail_addr).strip() and str(customer_name).strip() and str(mail_addr).count("@") ==1:
									if mail_addr.upper() not in alfa.keys():
										alfa[mail_addr.upper()]=[1,[document,sheet,str(row + 1),str(customer_name).strip()]+list(content_information[2:])]
									else:
										alfa[mail_addr.upper()].append([document,sheet,str(row+1),str(customer_name).strip()]+list(content_information[2:]))
										alfa[mail_addr.upper()][0] += 1
									#可视化处理：
									#showinfo("数据整理：" + mail_addr+"地址发现" + str(alfa[mail_addr][0]) + "个")
									mail_ok += 1
								else:
									empty_data.append([mail_addr,str(customer_name).strip(),document,sheet,str(row + 1)]+list(content_information[2:]))
									mail_ng += 1
							except Exception as e:
								showinfo("数据整理：整理异常，异常文件地址："+document+"-->"+sheet+"-->"+str(row) + "-->"+str(mail_addr)+"-->"+str(customer_name)+" 。报错信息："+ str(e))
					count_huoye += 1
			except Exception as e:
				showinfo("数据整理：整理异常，异常文件路径："+document+"-->"+sheet+""+" 。报错信息："+ str(e))
			count_file += 1
		showinfo("数据整理：Done. 总计处理{}个文件，{}个活页薄，{}个合格邮件地址，{}个异常邮件地址。".format(count_file,count_huoye,mail_ok,mail_ng))
		write_excel_1(empty_data)
		showinfo("结果输出：数据写入中......")
		write_excel_2(alfa)
		showinfo("结果输出：Done！")
	except Exception as e:
		showinfo("程序报错：data_deal  "+str(e))

def big_data_deal_file():
	#while True:
	filepaths = askopenfilenames(filetypes=[("office 2007 excel files","*.xls"),(" office 2010 excel files ","*.xlsx"),\
                                ("All files","*.*")], title="Select files")  # 选择打开什么文件，返回文件名
	filenames.set(str(len(filepaths)) + "个文件被选择")             # 设置变量filename的值
	showinfo('文件选择：'+'、'.join(filepaths))
	button4.config(state='disabled')
	data_deal(filepaths,mail_col_num,cus_col_num,data_row_num)
	button4.config(state='normal')

def feedback(title,msg,color='red'):
    newwindow=tkinter.Tk()
    newwindow.title(title)
    newwindow.minsize(300,100)
    newwindow.maxsize(300,100)
    label3 = tkinter.Label(newwindow,text=msg,justify='left')
    label3.configure(fg=color)
    label3.place(x=25,y=24)


Entry_CheckMail.bind("<Return>", checkemail)
Entry_CheckMail.configure(fg='blue') 

root.mainloop()